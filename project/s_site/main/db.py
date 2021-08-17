import sqlite3
import random
import time
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from helpTools import dump


def writeXLS(name,titles, data):
    filename = '%s.xlsx' % (name)
    wb = Workbook()
    ws = wb.active
    ws.append(titles)
    wb.save(filename)
    wb.close()

    wb = load_workbook(filename=filename, read_only=False)
    ws = wb.active
    for each in data:
        ws.append(each)
    wb.save(filename)
    wb.close()








def add_organization(id ,orgName ,orgCategory ,orgType ,latitude ,longitude ,address ,INN, urAddres ,okved ,ogrn ,profit ,gain ,compPrice ,site ,linkGis ,City ,phones):
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()
    myCursor.execute(f"SELECT id FROM orgs_List WHERE id ='{id}'")
    if myCursor.fetchone() is None:
        myCursor.execute(f"INSERT INTO orgs_List VALUES(? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,? ,?,?)", (id ,orgName ,orgCategory ,orgType ,float(latitude), float(longitude), address ,INN, urAddres ,okved ,ogrn ,profit ,gain ,compPrice ,site ,linkGis ,City ,phones))
        con.commit()
    else:
        print("не добавлен")
    myCursor.close()
    con.close()

def xlsxparser():
    data_name="eztnmzabix.xlsx"
    wb = load_workbook(data_name)
    ws = wb.active
    whichNow = 0
    for row in ws.iter_rows(min_row=0, max_row=92480, max_col=29):
        if whichNow==0:
            whichNow = whichNow + 1
            continue
        whichNow=whichNow+1
        id = " "
        orgName = " "
        orgCategory = " "
        orgType = " "
        latitude = " "
        longitude = " "
        address = " "
        INN = " "
        urAddres = " "
        okved = " "
        ogrn = " "
        profit = " "
        gain = " "
        compPrice = " "
        site = " "
        linkGis = " "
        City = " "
        phones = " "
        thisNumberOfRow = 0
        for cell in row:
            thisNumberOfRow = thisNumberOfRow+1
            if thisNumberOfRow == 1:
                orgName = str(cell.value)
            elif thisNumberOfRow == 2:
                orgType = str(cell.value)
            elif thisNumberOfRow == 4:
                INN = str(cell.value)
            elif thisNumberOfRow == 5:
                urAddres = str(cell.value)
            elif thisNumberOfRow == 9:
                City = str(cell.value)
            elif thisNumberOfRow == 10:
                address = str(cell.value)
            elif thisNumberOfRow == 11:
                orgCategory = str(cell.value)
            elif thisNumberOfRow == 13:
                phones = str(cell.value)
            elif thisNumberOfRow == 19:
                site = str(cell.value)
            elif thisNumberOfRow == 25:
                latitude = str(cell.value)
            elif thisNumberOfRow == 26:
                longitude = str(cell.value)
            elif thisNumberOfRow == 27:
                id = str(cell.value)
            elif thisNumberOfRow == 28:
                linkGis = str(cell.value)
            id = str(id)+str(whichNow)

        if latitude == "undefined" or latitude == "" or latitude == "None" or latitude == " " or latitude == None:
            latitude = 0.0
            print("undefined")
        if longitude == "undefined" or longitude == "" or longitude == "None" or longitude == " " or longitude == None:
            print("undefined")
            longitude = 0.0
        latitude = float(latitude)
        longitude = float(longitude)

        # try:
        #     latitude = float(latitude)
        #     longitude = float(longitude)
        # except Exception as e:
        #     print(latitude)
        #     print(longitude)
        #     print("неПередлан", e)

        # print(latitude)
        # print(longitude)
        # print(linkGis)

        print(str(whichNow))

        add_organization(id, orgName, orgCategory, orgType, latitude, longitude, address, INN, urAddres , okved, ogrn, profit, gain, compPrice, site, linkGis, City, phones)





def getAllORGS():
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()

    myCursor.execute("""SELECT urAddres from orgs_List""")
    items = myCursor.fetchall()
    myCursor.close()
    con.close()
    return items
    pass





def getTable():
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()
    resul = []
    iters = 0
    for each in myCursor.execute(f"SELECT * FROM orgs_List"):
        iters+=1
        if iters>10:
            break
        elif iters<5:
            continue
        resul.append(each)
    names = [description[0] for description in myCursor.description]
    myCursor.close()
    con.close()
    return names , resul



def plot (justDataOrListOfNamesForData, DataOptional = ["optional"]):
    from pylsy import pylsytable
    data = []
    names = []
    if DataOptional== ["optional"]:
        names = justDataOrListOfNamesForData[0]
        table = pylsytable(names)
        data=justDataOrListOfNamesForData[2:]
    else:
        names= justDataOrListOfNamesForData
        table = pylsytable(names)
        data = DataOptional

    for each in data:
        indexForIterator = 0
        for any in each:
            table.append_data(names[indexForIterator], str(any))
            indexForIterator+=1

    print(table)
    dump("folder", "file", table)


# s , ss = getTable()
# data = ss.insert(0,s)
# writeXLS("db",s, ss)
# plot(s,ss)




def getOrgsSquare(latMin,latMax, lonMin,lonMax):
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()

    myCursor.execute(f"SELECT * FROM orgs_List WHERE  longitude  BETWEEN {lonMin} AND {lonMax};")
    myCursor.execute(f"SELECT * FROM orgs_List WHERE  latitude BETWEEN {latMin} AND {latMax};")


    orgsInSquare = myCursor.fetchall()
    myCursor.close()
    con.close()
    print(orgsInSquare)
    return orgsInSquare
    pass



def update_org(INN,compPrice,gain,profit):
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()
    myCursor.execute(f"UPDATE orgs_List SET compPrice ='{compPrice}' WHERE INN ='{INN}' ")
    myCursor.execute(f"UPDATE orgs_List SET gain ='{gain}' WHERE INN ='{INN}' ")
    myCursor.execute(f"UPDATE orgs_List SET profit ='{profit}' WHERE INN ='{INN}' ")
    con.commit()
    myCursor.close()
    con.close()
    pass

print(len(getOrgsSquare(48.6812,48.7279,44.4557,44.5905)))
# update_org(3443009921,compPrice,gain,profit)



if __name__ == '__main__':
    # timestart = get_time
    # print(timestart)


    print("База данных импортирована")
    con = sqlite3.connect('ORGSSM.db')
    myCursor = con.cursor()

    myCursor.execute("""CREATE TABLE IF NOT EXISTS orgs_List (
    id TEXT,
    orgName TEXT,
    orgCategory TEXT,
    orgType TEXT,
    latitude REAL,
    longitude REAL,
    address TEXT,
    INN TEXT,
    urAddres TEXT,
    okved TEXT,
    ogrn TEXT,
    profit TEXT,
    gain TEXT,
    compPrice TEXT,
    site TEXT,
    linkGis TEXT,
    City TEXT,
    phones TEXT
    )""")

    con.commit()
    myCursor.close()
    con.close()
    # xlsxparser()
    # print(getAllORGS())
    # print(getTable())
    # plot()



    # print(getAllORGS())
    # timeend = get_time
    # print("timestart",timestart)
    # print("timeend",timeend)














# def waitUsersGet():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     users = []
#     for each in myCursor.execute(f"SELECT * FROM KeyWaitUsers"):
#         # print(each[0])
#         # print(each)
#         users.append(each)
#     # print("losts",lostS)
#     myCursor.close()
#     con.close()
#     return users
#
# def waitUsersadd(id):
#     print("waitUsersadd",id)
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"SELECT id_ FROM KeyWaitUsers WHERE id_ ='{id}'")
#     if myCursor.fetchone() is None:
#         myCursor.execute(f"INSERT INTO KeyWaitUsers VALUES (?)", (id,))
#         con.commit()
#     else:
#         print("уже в списке ожидающих")
#         # for each in myCursor.execute("SELECT * FROM subscribers"):
#         #     print(each)
#     myCursor.close()
#     con.close()
#     pass
#
# def waitUsersremove(id):
#     print("waitUsersremove",id)
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"DELETE FROM KeyWaitUsers WHERE id_ ='{id}' ")
#     con.commit()
#     myCursor.close()
#     con.close()
#     pass
# def waitForPostRead():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute("""SELECT booler FROM waitForPost Where  id_ = 0 """)
#     waiters = myCursor.fetchone()
#     myCursor.close()
#     con.close()
#     return waiters[0]
#     pass
# def waitForPostChange(booler):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"UPDATE waitForPost SET booler ='{booler}' WHERE id_ =0 ")
#     con.commit()
#     myCursor.close()
#     con.close()
#     pass
# def replyToSubersRead():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute("""SELECT booler FROM replyToSubers Where  id_ = 0 """)
#     waiters = myCursor.fetchone()
#     myCursor.close()
#     con.close()
#     return waiters[0]
#     pass
# def replyToSubersChange(booler):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"UPDATE replyToSubers SET booler ='{booler}' WHERE id_ =0 ")
#     con.commit()
#     myCursor.close()
#     con.close()
#     pass
#
#
#
# def replyToDeadSubersRead():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute("""SELECT booler FROM replyToDeadSubers Where  id_ = 0 """)
#     waiters = myCursor.fetchone()
#     myCursor.close()
#     con.close()
#     return waiters[0]
#     pass
#
# def replyToDeadSubersChange(booler):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"UPDATE replyToDeadSubers SET booler ='{booler}' WHERE id_ =0 ")
#     con.commit()
#     myCursor.close()
#     con.close()
#     pass








#
# def unsubscribeONE(id):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#
#
#     user = None
#     for each in myCursor.execute(f"SELECT * FROM subscribers WHERE id ='{id}'"):
#         user=each
#     myCursor.execute(f"SELECT id FROM dead_subscribers WHERE id ='{id}'")
#     if myCursor.fetchone() is None:
#         myCursor.execute(f"INSERT INTO dead_subscribers VALUES(?,?,?,?,?,?)", (user))
#     myCursor.execute(f"DELETE FROM subscribers WHERE id ='{id}' ")
#     con.commit()
#     myCursor.close()
#     con.close()
# unsubscribeONE(1006130578)
#
# def unsubscribeREGULAR(now):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#
#     lostS = []
#     for each in myCursor.execute(f"SELECT * FROM subscribers WHERE end <='{now}'"):
#         # print(each[0])
#         # print(each)
#         lostS.append(each)
#     for rrr in lostS:
#         myCursor.execute(f"INSERT INTO dead_subscribers VALUES(?,?,?,?,?,?)", (rrr))
#         myCursor.execute(f"DELETE FROM subscribers WHERE id ='{rrr[0]}' ")
#         con.commit()
#     # print("losts",lostS)
#     con.commit()
#     myCursor.close()
#     con.close()
#     return lostS
#
#
#
# def CreateKey(days):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#
#     # with sql.connect("database.db") as con:
#     #     myCursor = con.cursor()
#     def get_random_key():
#         def r():
#             return random.randint(42, 4242)
#         return r() * r() + r()
#     new_key = get_random_key()
#     myCursor.execute(f"SELECT * FROM keys WHERE key ='{new_key}'")
#     if myCursor.fetchone() is None:
#         myCursor.execute(f"INSERT INTO keys VALUES(?,?)", (new_key, days*24*60*60))#2678400 = 1 month
#         con.commit()
#         # print("ключ",new_key," добавлен")
#         pass
#     else:
#         print("такой ключ уже есть")
#
#     # myCursor.close()
#     # con.close()
#
#     myCursor.close()
#     con.close()
#     return new_key
#     pass
#
#
#
#
# def CreateKeys(days,howMany):
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     keyList = []
#
#     def get_random_key():
#         return random.randint(42, 4242424242) + random.randint(42, 4242424242)
#
#     for each in range(howMany):
#         new_key = get_random_key()
#         myCursor.execute(f"SELECT * FROM keys WHERE key ='{new_key}'")
#         if myCursor.fetchone() is None:
#             myCursor.execute(f"INSERT INTO keys VALUES(?,?)", (new_key, days*24*60*60))#2678400 = 1 month
#             keyList.append(new_key)
#             # print("ключ",new_key," добавлен")
#             pass
#         else:
#             print("такой ключ уже есть")
#
#     con.commit()
#     myCursor.close()
#     con.close()
#     return keyList
#     pass
#
#
# def ActivateKey(id, first_name,username,start, key):#этот модуль надо будет жестко тестить
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"SELECT time FROM keys WHERE key ='{key}'")
#     tt = myCursor.fetchone()[0]
#     print("tt",tt)
#     end = start+tt
#     his_time = tt/60/60/24
#     print(end)
#     relevance = True
#     myCursor.execute(f"SELECT * FROM subscribers WHERE id ='{id}'")
#     if myCursor.fetchone() is None:
#         myCursor.execute(f"DELETE FROM keys WHERE key ='{key}' ")
#
#
#         myCursor.execute(f"SELECT id FROM subscribers WHERE id ='{id}'")
#         if myCursor.fetchone() is None:
#             myCursor.execute(f"INSERT INTO subscribers VALUES(?,?,?,?,?,?)",
#                              (id, first_name, username, start, end, relevance))
#             print("подписчик ", first_name, " ", username, " добавлен в список подписчиков")
#             myCursor.execute(f"DELETE FROM dead_subscribers WHERE id ='{id}' ")
#             con.commit()
#
#         else:
#
#             print("подписка ещё активна")
#
#
#
#         con.commit()
#     else:
#         print("____________________else")
#         myCursor.execute(f"SELECT end FROM subscribers WHERE id ='{id}'")
#         user_end = myCursor.fetchone()[0]
#
#         print("user_end",user_end)
#
#         # print("key time ", tt)
#         new_end = int(user_end)+int(tt)
#         his_time = (new_end - start)/60/60/24
#         print("new_end", new_end)
#         myCursor.execute(f"UPDATE subscribers SET end ='{new_end}' WHERE     id ='{id}'")
#
#         myCursor.execute(f"DELETE FROM keys WHERE key ='{key}' ")
#         con.commit()
#     ss = 'Спасибо, ваш ключ активирован, подписка закончится через '+ str(his_time)+" день"
#     print(ss)
#     myCursor.close()
#     con.close()
#     return ss
#     pass
#
#
#
#
#
# def getAllSubs():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#     myCursor.execute("""SELECT * from subscribers""")
#     items = myCursor.fetchall()
#     myCursor.close()
#     con.close()
#     return items
#     pass
#
# def getAllDEADSubs():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#     myCursor.execute("""SELECT * from dead_subscribers""")
#     items = myCursor.fetchall()
#
#     myCursor.close()
#     con.close()
#     return items
#     pass
#
#
#
# def printall():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#
#     print("keys")
#     for each in myCursor.execute("SELECT * FROM keys"):
#         print(each)
#     print("subscribers")
#     for each in myCursor.execute("SELECT * FROM subscribers"):
#         print(each)
#     print("dead_subscribers")
#     for each in myCursor.execute("SELECT * FROM dead_subscribers"):
#         print(each)
#     print("post_que")
#     for each in myCursor.execute("SELECT * FROM post_que"):
#         print(each)
#     print(waitUsersGet())
#     print("replyToSubersRead", replyToSubersRead())
#     print("waitForPostRead", waitForPostRead())
#     print("replyToDeadSubers", replyToDeadSubersRead())
#
#     myCursor.close()
#     con.close()
#
#
# def getAllPosts():
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute("""SELECT * from post_que ORDER BY time""")
#     posts = myCursor.fetchall()
#     myCursor.close()
#     con.close()
#     # print(posts)
#     return posts
#
# def addPost(time_, text_,photos,audio):
#     time_ = int(time_)
#     con = sqlite3.connect('teleBot.db')
#     myCursor = con.cursor()
#     myCursor.execute(f"SELECT * FROM post_que WHERE time ='{time_}'")
#     if myCursor.fetchone() is None:
#         print("myCursor.fetchone()---------------------",myCursor.fetchone())
#         myCursor.execute(f"INSERT INTO post_que VALUES(?,?,?,?)", (time_, text_,photos,audio,))
#         con.commit()
#         print("пост добавлен\n")
#         pass
#     else:
#         print("такой пост уже есть")
#         # for each in myCursor.execute("SELECT * FROM post_que"):
#         #     print(each)
#     myCursor.close()
#     con.close()
#
#
#
#
# def del_post (number):
#     try:
#         number = int(number)
#         con = sqlite3.connect('teleBot.db')
#         myCursor = con.cursor()
#         myCursor.execute("""SELECT * FROM post_que ORDER BY time""")
#         post = myCursor.fetchall()
#         print(post)
#         print("-----")
#         number = int(post[number][0])
#         print(number)
#         myCursor.execute(f"SELECT * FROM post_que WHERE time ='{number}'")
#         post1 = myCursor.fetchone()
#         myCursor.execute(f"DELETE FROM post_que WHERE time ='{number}' ")
#         con.commit()
#         myCursor.close()
#         con.close()
#         return post1
#     except:
#         return "не смог удалить пост. возможно вы указали номер поста, которого нет. Либо список постов пуст"
#
#
#
# def DeleteAndGetPost():
#         con = sqlite3.connect('teleBot.db')
#         myCursor = con.cursor()
#         myCursor.execute("""SELECT MIN(time) from post_que""")
#         post = int(myCursor.fetchone()[0])
#
#         myCursor.execute(f"SELECT * FROM post_que WHERE time ='{post}'")
#         post1 = myCursor.fetchone()
#         myCursor.execute(f"DELETE FROM post_que WHERE time ='{post}' ")
#         con.commit()
#         myCursor.close()
#         con.close()
#         return post1





